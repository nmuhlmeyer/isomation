'''**********************************************************************


    This script reads an ADS exported *.csv file and creates pivot tables
    based on the exported data.
    Copyright (C) 2016  Nicholas Muhlmeyer

    This program is free software: you can redistribute it and/or modify
    it under the terms of the GNU General Public License as published by
    the Free Software Foundation, either version 3 of the License, or
    (at your option) any later version.

    This program is distributed in the hope that it will be useful,
    but WITHOUT ANY WARRANTY; without even the implied warranty of
    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
    GNU General Public License for more details.

    You should have received a copy of the GNU General Public License
    along with this program.  If not, see <http://www.gnu.org/licenses/>.
**********************************************************************'''
import re
import pandas
import numpy
import csv
import os.path
import openpyxl

def main():
    #filename = raw_input('Select a file: ')
    filename = 'matrix'
    FullFilename = filename + '.csv'
    isoMatrix = extractIsolation(FullFilename)
    #isoMatrix = extractIsolation2(FullFilename)

    #createSpec(isoMatrix,filename)
    df = pandas.DataFrame(isoMatrix,columns=['Port 1','Port 2','Isolation'])
    dfFullPivot = df.pivot('Port 1', 'Port 2')
    realFile = os.path.isfile(filename + 'Info_1.csv')
    subset = 1
    while realFile:
        NewFilename = filename + 'Sub_' + str(subset) + '.csv'
        rowdata=[]
        with open(filename + 'Info_' + str(subset) + '.csv') as csvfile:
            inF = csv.reader(csvfile,delimiter=',')
            for row in inF:
                rowdata.append(row)
        desiredRow=[int(n) for n in rowdata[0]]
        desiredColumn=[int(n) for n in rowdata[1]]
        desiredRowSorted, desiredColumnSorted = subsetOrdering(isoMatrix,desiredRow,desiredColumn)
        dfTrim = pivotSubset(dfFullPivot,desiredRowSorted,desiredColumnSorted)
        subset += 1
        InfoFile = filename + 'Info_' + str(subset) + '.csv'        
        realFile = os.path.isfile(InfoFile)
        #print(dfTrim)
        dfTrim.to_csv(NewFilename)
        #createSpec(isoMatrix,filename)
    #print dfFullPivot
    dfFullPivot.to_csv(filename + 'Full.csv')
    return

'''**************************************************'''

def extractIsolation(FullFilename):
    portNameDict, _ = portNameDictionary()
    with open(FullFilename) as inF:
        points = 0
        for index,line in enumerate(inF):
            desired_line = re.match('Num. Points : \D',line,re.M|re.I)
            if desired_line:
                desired_line2 = re.search(']',line,re.M|re.I)
                if desired_line2: points = line[desired_line.end():desired_line2.start()]
    points = int(points)
    isoMatrix = init_2dlist(points,3)
    i = 0
    with open(FullFilename) as csvFile:
        inF = csv.reader(csvFile,delimiter=',')
        for row in inF:
            if i > 0:
                isoMatrix[i-1][0] = portNameDict[row[0][1:]]                
                isoMatrix[i-1][1] = portNameDict[row[1][:-1]]
                isoMatrix[i-1][2] = int(round(float(row[-1]),0))
            i += 1
    return isoMatrix

def portNameDictionary():
    filename = 'portName.csv'
    with open(filename) as csvFile:
        line = csv.reader(csvFile)
        portInfo = list(line)
        numberPorts = numpy.size(portInfo)/len(portInfo)
        highestPort = int(portInfo[-1][-1])
        portNameDict = {}
        for j in range(numberPorts):
            portNameDict[portInfo[1][j]] = portInfo[0][j]
    return portNameDict, highestPort

def pivotSubset(dfFullPivot,desiredRow,desiredColumn):
    indexRow = init_count_list(len(dfFullPivot))
    indexColumn = init_count_list(len(dfFullPivot))
    [indexRow.pop(k) for k in sorted(desiredRow,reverse=True)]
    [indexColumn.pop(k) for k in sorted(desiredColumn,reverse=True)]
    dfTrim = dfFullPivot.drop(dfFullPivot.index[indexRow]) # to get desired rows
    dfTrim = dfTrim.transpose() #to be abel to edit columns
    dfTrim = dfTrim.drop(dfTrim.index[indexColumn]) # to get desired columns
    dfTrim = dfTrim.transpose()
    return dfTrim

def subsetOrdering(isoMatrix,desiredRow,desiredColumn):
    portNameDict, highestPort = portNameDictionary()
    desiredRowName = init_list(len(desiredRow))
    desiredColumnName = init_list(len(desiredColumn))
    portOrder=[]
    i = 0
    for j in range(highestPort):
        i += 1
        if str(j+1) in portNameDict.keys(): 
            i += -1
            portOrder.append(portNameDict[str(j+1)])
    for j in range(len(desiredRowName)): desiredRowName[j] = portNameDict[str(desiredRow[j])]
    for j in range(len(desiredColumnName)): desiredColumnName[j] = portNameDict[str(desiredColumn[j])]
    portOrderSorted = sorted(portOrder)
    desiredRowSorted = []
    for port in desiredRowName: desiredRowSorted.append(portOrderSorted.index(port))
    desiredColumnSorted = []
    for port in desiredColumnName: desiredColumnSorted.append(portOrderSorted.index(port))
    return desiredRowSorted, desiredColumnSorted

def init_2dlist(num_rows,num_cols):
    return [[0 for j in range(num_cols)] for i in range(num_rows)]
def init_list(num_rows):
    return [0 for i in range(num_rows)]
def init_count_list(num_elements):
    return [k for k in range(num_elements)]

#main()
'''**************************************************'''

def inverse_factorial(x):
    if x < 1: return 'Error: x less than 1'
    y = 0
    while x > 1:
        y = y + 1
        x = x/y
    if y == 0: return 1
    return y

def CSVtoXLS(filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    with open(filename, 'rb') as inF:
        reader = csv.reader(inF)
        for r, row in enumerate(reader, start = 1):
            for c, val in enumerate(row, start = 1):
                ws.cell(row = r, column = c).value = val
    #outOfSpec(wb,ws,r,c)
    wb.save(filename[:-4] + '.xlsx')
    os.remove(filename)
    return

def outOfSpec(wb,ws,r,c,wbSpec,wsSpec):
    badSpecFont = openpyxl.styles.Font(color='FFFF0000') # RED
    r += -1
    c += -1
    #a = 5 # row (number)
    #b = 7 # column (letter)
    #spec = -30
    print xlref(a,b)
    if int(round(float(ws[xlref(a,b)].value))) > int(round(float(wsSpec[xlref(a,b)].value))): ws[xlref(a,b)].font = badSpecFont # mark out of spec cell
    #if int(round(float(ws[xlref(a,b)].value))) > spec: ws[xlref(a,b)].font = badSpecFont # mark out of spec cell
    return

def xlref(row, column, zero_indexed=True):
    if zero_indexed:
        row += 1
        column += 1
    return openpyxl.utils.get_column_letter(column) + str(row)

def createSpec(isoMatrix,filename):
    specFile = filename + 'Spec.csv'
    specFileXls = filename + 'Spec.xlsx'
    if os.path.isfile(specFile):
        #a = 1
        #wbSpec = openpyxl.Workbook(specFile[:-4] + '.xlsx')
        #wsSpec = wb.active
        #outOfSpec(wb,ws,r,c,wbSpec,wsSpec)
        
        '''with open(specFile) as csvFile:
            inF = csv.reader(csvFile,delimiter=',')
            for row in inF:
                print row'''
        #xlsCreator(filename)
    else: #create template and write xls file
        isoSpec = isoMatrix
        for i in range(len(isoSpec)): isoSpec[i][2] = int(0)
        dfSpec = pandas.DataFrame(isoMatrix,columns=['Port 1','Port 2','Isolation'])
        dfSpec = dfSpec.pivot('Port 1', 'Port 2')
        dfSpec.to_csv(specFile)
        CSVtoXLS(specFile)
        print 'Please udpated the Isolation Spec matrix ' + specFileXls
    return

def extractIsolation2(FullFilename):
    portNameDict, _ = portNameDictionary()
    with open(FullFilename) as inF:
        points = 0
        for index,line in enumerate(inF):
            desired_line = re.match('Num. Points : \D',line,re.M|re.I)
            if desired_line:
                desired_line2 = re.search(']',line,re.M|re.I)
                if desired_line2: points = line[desired_line.end():desired_line2.start()]
    points = int(points)
    i = 0
    if os.path.isfile('specMatrix' + '.csv'):
        isoMatrix = init_2dlist(points,4)
        specValue = readSpec('specMatrix.csv')
        with open('isoSpecTable.csv','wb') as csvOut:
            csvwriter = csv.writer(csvOut, delimiter=',')
            with open(FullFilename) as csvFile:
                inF = csv.reader(csvFile,delimiter=',')
                for row in inF:
                    if i > 0:
                        isoMatrix[i-1][0] = portNameDict[row[0][1:]]
                        isoMatrix[i-1][1] = portNameDict[row[1][:-1]]
                        isoMatrix[i-1][2] = specValue[i-1]
                        isoMatrix[i-1][3] = int(round(float(row[-1]),0))
                        csvwriter.writerow([isoMatrix[i-1][0], isoMatrix[i-1][1], isoMatrix[i-1][2], isoMatrix[i-1][3]])
                    i += 1
    else:
        isoMatrix = init_2dlist(points,3)
        with open('specMatrix' + '.csv','wb') as csvOut:
            csvwriter = csv.writer(csvOut, delimiter=',')
            with open(FullFilename) as csvFile:
                inF = csv.reader(csvFile,delimiter=',')
                for row in inF:
                    if i > 0:
                        isoMatrix[i-1][0] = portNameDict[row[0][1:]]
                        isoMatrix[i-1][1] = portNameDict[row[1][:-1]]
                        isoMatrix[i-1][2] = 0
                        csvwriter.writerow([isoMatrix[i-1][0], isoMatrix[i-1][1], isoMatrix[i-1][2]])
                    i += 1
    return isoMatrix

'''def createSpec2(points):
    portNameDict, _ = portNameDictionary()
    specMatrix = init_2dlist(points,3)
    i = 0
    print portNameDict[[1][1:]]
    with open('eggs2.csv','wb') as csvFile:
        csvWriter = csv.writer(csvFile, delimiter=',')
        for i in range(len(specMatrix)):
            if i > 0:
                specMatrix[i-1][0] = 1
                specMatrix[i-1][1] = 2
                specMatrix[i-1][2] = 0
                csvWriter.writerow([specMatrix[i-1][0],specMatrix[i-1][1],specMatrix[i-1][2]])
            i += 1
    return
'''

def readSpec(filename):
    specValue = []
    with open(filename) as inF:
        inF = csv.reader(inF,delimiter=',')
        for row in inF:
            specValue.append(int(round(float(row[-1]),0)))
    return specValue

def outOfSpec2(wb,ws,r,c,wbSpec,wsSpec):
    badSpecFont = openpyxl.styles.Font(color='FFFF0000') # RED
    r += -1
    c += -1
    #a = 5 # row (number)
    #b = 7 # column (letter)
    #spec = -30
    print xlref(a,b)
    if int(round(float(ws[xlref(a,b)].value))) > int(round(float(ws[xlref(a,b-1)].value))): ws[xlref(a,b)].font = badSpecFont # mark out of spec cell
    #if int(round(float(ws[xlref(a,b)].value))) > spec: ws[xlref(a,b)].font = badSpecFont # mark out of spec cell
    return

main()